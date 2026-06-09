from datetime import datetime
from pathlib import Path
import html
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _agora_formatado():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def _sufixo_arquivo():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _pasta_downloads():
    pasta = Path.home() / "Downloads"
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta


def _texto_filtros_html(filtros):
    return "".join(
        f"<li><strong>{html.escape(str(chave))}:</strong> "
        f"{html.escape(str(valor or 'Todos'))}</li>"
        for chave, valor in filtros.items()
    )


def _linhas_html(linhas):
    resultado = []
    for linha in linhas:
        colunas_html = "".join(
            f"<td>{html.escape(str(valor or ''))}</td>" for valor in linha
        )
        resultado.append(f"<tr>{colunas_html}</tr>")
    return "".join(resultado)


def gerar_relatorio_html(filtros, cabecalhos, linhas, titulo="Relatório do Painel do Robô"):
    cabecalho_html = "".join(
        f"<th>{html.escape(str(cabecalho))}</th>" for cabecalho in cabecalhos
    )
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>{html.escape(titulo)}</title>
<style>
@page {{ size: A4 landscape; margin: 10mm; }}
body {{ font-family: Arial, sans-serif; margin: 18px; color: #1f2937; }}
.barra-acoes {{ display: flex; justify-content: flex-end; margin-bottom: 14px; }}
.btn-imprimir {{ background: #1f538d; color: white; border: none; border-radius: 8px; padding: 10px 16px; font-weight: bold; cursor: pointer; }}
h1 {{ color: #1f538d; margin-bottom: 4px; }}
.meta {{ margin-bottom: 16px; color: #4b5563; }}
.filtros {{ background: #f3f4f6; border: 1px solid #d1d5db; border-radius: 8px; padding: 12px 16px; margin-bottom: 18px; }}
table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
th, td {{ border: 1px solid #d1d5db; padding: 6px; font-size: 11px; vertical-align: top; word-break: break-word; }}
th {{ background: #1f538d; color: white; }}
tr:nth-child(even) {{ background: #f9fafb; }}
@media print {{ body {{ margin: 0; }} .barra-acoes {{ display: none; }} }}
</style>
</head>
<body>
<div class="barra-acoes"><button class="btn-imprimir" onclick="window.print()">Imprimir PDF A4</button></div>
<h1>{html.escape(titulo)}</h1>
<div class="meta">Gerado em {html.escape(_agora_formatado())} | Total de notas: {len(linhas)}</div>
<div class="filtros"><strong>Filtros aplicados</strong><ul>{_texto_filtros_html(filtros)}</ul></div>
<table><thead><tr>{cabecalho_html}</tr></thead><tbody>{_linhas_html(linhas)}</tbody></table>
</body>
</html>"""


def salvar_relatorio_excel(filtros, cabecalhos, linhas, caminho_saida=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Filtradas"
    ws["A1"] = "Relatório do Painel do Robô"
    ws["A1"].font = Font(size=14, bold=True, color="1F538D")
    ws["A2"] = f"Gerado em {_agora_formatado()}"
    ws["A3"] = f"Total de notas: {len(linhas)}"
    linha_atual = 5
    ws.cell(row=linha_atual, column=1, value="Filtros aplicados").font = Font(bold=True)
    linha_atual += 1
    for chave, valor in filtros.items():
        ws.cell(row=linha_atual, column=1, value=str(chave))
        ws.cell(row=linha_atual, column=2, value=str(valor or "Todos"))
        linha_atual += 1
    linha_atual += 1
    header_fill = PatternFill(fill_type="solid", start_color="1F538D", end_color="1F538D")
    for idx, cabecalho in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=linha_atual, column=idx, value=str(cabecalho))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for linha in linhas:
        linha_atual += 1
        for idx, valor in enumerate(linha, start=1):
            ws.cell(row=linha_atual, column=idx, value=str(valor or "")).alignment = Alignment(vertical="top", wrap_text=True)
    caminho = Path(caminho_saida) if caminho_saida else (
        _pasta_downloads() / f"relatorio_painel_robo_{_sufixo_arquivo()}.xlsx"
    )
    wb.save(caminho)
    return caminho
